# Databricks notebook source
# MAGIC %pip install azure-storage-blob azure-identity azure-core openai exception python-docx azure-ai-formrecognizer openpyxl tabulate azure-cosmos
# MAGIC

# COMMAND ----------

dbutils.widgets.text("anio", "2024")

# COMMAND ----------

anio = dbutils.widgets.get("anio").split("/")[-1]
anio

# COMMAND ----------

import time
import os
import io
from io import BytesIO
from azure.core.credentials import AzureKeyCredential
from azure.storage.blob import BlobServiceClient
from openai import AzureOpenAI
import pandas as pd
import json
from datetime import datetime
import docx
from azure.ai.formrecognizer import DocumentAnalysisClient
from openpyxl import load_workbook
from tabulate import tabulate
from azure.cosmos import CosmosClient

# COMMAND ----------

#region Variables
removeall=False
remove=False
localpdfparser=False
storageaccount= 'asapocifiscr'
credential = 'HCqYUM++aT/vvrrWq3g9WDkqcHEredVPD1ujqtAmf7wfNCS5crVge+8pY0mjieil4Dbs4gdyHtCW+AStd0gcfw=='
formrecognizerservice='di-POC-IFIS-CR'
credentialformrecognizer ='CgvcBr35Hikf05T6vZsQPnysCHFlOo8S0dWP46o3ST1yraReK5rIJQQJ99AKACYeBjFXJ3w3AAALACOGS8Jv'
source_container_name = 'source'
target_container_name = 'source'
verbose=False
AZURE_OPENAI_SERVICE = "oai-POC-iDataFactory-CR-IFISCAF"
OPENAI_API_KEY = "7c4vG9I7g2YVz8bL1kbkSVBkCo5yCXiMnAp6XQGgR4aAnyuZuPFVJQQJ99AKACYeBjFXJ3w3AAABACOGA3HD"
AZURE_OPENAI_ENDPOINT= "https://oai-poc-idatafactory-cr-ifiscaf.openai.azure.com/"


#Extraccion datos S&P
client = AzureOpenAI(
    api_key=OPENAI_API_KEY,  
    api_version="2024-10-21",
    azure_endpoint = AZURE_OPENAI_ENDPOINT
    )
blob_service = BlobServiceClient(
    account_url=f"https://{storageaccount}.blob.core.windows.net",
    credential=credential
)
source_container_client = blob_service.get_container_client(source_container_name)



lista_prompt_extraccion = [ 
f"""A partir de los datos proporcionados en el contexto, realiza una extracción de información siguiendo estrictamente estas instrucciones:

1Usa únicamente la información dada, sin agregar datos externos.
Ten en cuenta que las variables pueden estar distribuidas por diferentes años de análisis; identifica y organiza los datos según el año correspondiente.
La respuesta debe estar en español, en formato JSON, clara y concisa, utilizando siempre comillas dobles.
Los nombres de los campos en el JSON deben tener un guion bajo "_" en los espacios.
Si tienes que listar datos, usa el símbolo "|" como separador en lugar de una coma ",".
Solo responde con el JSON, sin saltos de línea ni texto adicional, y sin la palabra "JSON" al inicio o al final.
Asegúrate de que el JSON tenga el formato correcto.
Los campos por extraer están en un array,cada uno con un nombre de campo y una descripción.
Para los campos que incluyen montos, convierte el monto a un valor numérico sin unidades, si el monto está en formato como 'USD 60 M', convierte el valor a su representación numérica en base 10 por ejemplo USD 60 M a 60000000. 
Al brindar la información extraída ten en cuenta el tipo de dato que se indica en la descripción de cada campo.
"""
]



def _spreadsheet_process(filepath):
    blob_client = source_container_client.get_blob_client(filepath)
    data = blob_client.download_blob().readall()
    blob_stream = BytesIO(data)                   
    workbook = load_workbook(blob_stream, data_only=True)

    # Process each sheet in the workbook
    sheets = []
    
    for sheet_name in workbook.sheetnames:
        sheet_dict = {}            
        sheet_dict['name'] = sheet_name
        sheet = workbook[sheet_name]
        
        table = _excel_to_html(sheet)
        table = _excel_to_markdown(sheet)
        sheet_dict["table"] = table
        sheets.append(sheet_dict)
    
    return sheets

def _excel_to_markdown(sheet):
    # Read the data and determine cell colors
    data = []
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            cell_value = cell.value
            if cell_value is None:
                cell_value = ""
            cell_text = str(cell_value)
            row_data.append(cell_text)
        if "".join(row_data)!="":
            data.append(row_data)

    # Get the header from the first row
    headers = [cell.value if cell.value is not None else "" for cell in sheet[1]]
    table = tabulate(data, headers, tablefmt="pipe")
    return table

def _excel_to_html(sheet):
    html = '<table border="1">'
    
    # Dictionary to track merged cells
    merged_cells = {}
    
    # Process merged cells to map them to colspan and rowspan
    for merged_cell in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_cell.min_col, merged_cell.min_row, merged_cell.max_col, merged_cell.max_row
        merged_cells[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
    
    # Iterate over rows and columns to build the HTML
    for row in sheet.iter_rows():
        html += '  <tr>'
        for cell in row:
            row_num = cell.row
            col_num = cell.column
            
            # Check if the cell is the top-left of a merged cell
            if (row_num, col_num) in merged_cells:
                rowspan, colspan = merged_cells[(row_num, col_num)]
                cell_value = '' if cell.value is None else cell.value
                html += f'    <td rowspan="{rowspan}" colspan="{colspan}">{cell_value}</td>'
            else:
                # Skip cells that are part of a merged range but not the top-left
                is_merged = False
                for key, (rspan, cspan) in merged_cells.items():
                    start_row, start_col = key
                    if start_row <= row_num < start_row + rspan and start_col <= col_num < start_col + cspan:
                        is_merged = True
                        break
                
                if not is_merged:
                    cell_value = '' if cell.value is None else cell.value
                    html += f'    <td>{cell_value}</td>'
                
        html += '  </tr>'
    
    html += '</table>'
    html = html.replace('\n', '').replace('\t', '')
    return html

def get_document_text(file_path, blob_service, source_container_client, formrecognizerservice, credentialformrecognizer):
    # Descargar el archivo desde Azure Blob Storage
    blob_client = source_container_client.get_blob_client(file_path)
    file_stream = io.BytesIO()
    file_stream.write(blob_client.download_blob().readall())
    file_stream.seek(0)
    print(f"Loaded {file_stream.getbuffer().nbytes} bytes from {file_path} in Blob Storage.")

    # Detectar el tipo de archivo basado en la extensión
    file_extension = os.path.splitext(file_path)[1].lower()
    output_txt_filename = f"{os.path.splitext(file_path)[0]}.txt"

    if file_extension == '.pdf':
        # Usar Azure Form Recognizer para extraer texto de archivos PDF
        form_recognizer_client = DocumentAnalysisClient(
            endpoint=f"https://{formrecognizerservice}.cognitiveservices.azure.com/",
            credential=AzureKeyCredential(credentialformrecognizer)
        )
        
        poller = form_recognizer_client.begin_analyze_document("prebuilt-read", file_stream)
        result = poller.result()

        # Procesar resultados del PDF
        document_text = ""
        for page in result.pages:
            document_text += f"--- Page {page.page_number} ---\n"
            for line in page.lines:
                document_text += f"{line.content}\n"


    elif file_extension in ['.xls', '.xlsx']:
        # Procesar el archivo Excel
        df = pd.read_excel(file_stream)
        document_text = df.to_string(index=False)  # Convertir el DataFrame a texto

    else:
        raise ValueError(f"Formato de archivo no soportado: {file_extension}")

    print(f"Saved extracted text to {output_txt_filename}")
    return document_text

def getdata(filename):
    try:
        # Accede al cliente del blob
        proceseddocuments = source_container_client.get_blob_client(f"OUTPUT/{filename}")

        if proceseddocuments.exists():
            # Descargar el blob existente como bytes
            blob_data = proceseddocuments.download_blob().content_as_bytes()
            
            # Convertir los bytes a un DataFrame usando BytesIO
            proceseddocuments_df = pd.read_csv(BytesIO(blob_data))
            
            # Devolver el DataFrame
            return proceseddocuments_df
        else:
            # Si el blob no existe
            return {"result": "Error: File not found"}
    except Exception as e:
        # Manejo de errores
        return {"result": f"Error: {str(e)}"}



# COMMAND ----------



def ResultadoPais(filepathexcel,filepathsyp):

    print("Obteninedo Excel")
    resultexcel = _spreadsheet_process(filepathexcel)
    markdown = resultexcel[0]["table"]
    print("Obteninedo syp") 
    data_syp=get_document_text(
        file_path=filepathsyp,
        blob_service=blob_service,
        source_container_client=source_container_client,
        formrecognizerservice=formrecognizerservice,
        credentialformrecognizer=credentialformrecognizer
    )

    jsonfinal = """
        {
                    "PAIS": "Argentina",
                    "RATING": "BBB+",
                    "OUTLOOK": "POSTIIVO",
                    "VARPBI": [
                        123,
                        1345
                    ],
                    "VARPBIFECHA": [
                        "2023",
                        "2024"
                    ],
                    "INFLACIÓN": [
                        123,
                        1345
                    ],
                    "INFLACIÓNFECHA": [
                        "2023",
                        "2024"
                    ],
                    "BICAGROUP": 9,
                    "Economicresilience": "High risk",
                    "EconomicImbalance": "High risk",
                    "CreditRisk": "High risk",
                    "CreditRiskTrend": "High risk",
                    "Economicrisk": 9,
                    "GovermentSupport": "12434",
                    "industryrisk": 10,
                    "InstitutionalFramework": "12434",
                    "CompetitiveDynamics": "12434",
                    "SystemWideFunding": "12434",
                    "IndustryRiskTrend": "12434",
                    "Contexto": "En el año más reciente, Argentina muestra una situación económica y financiera compleja. La inflación ha alcanzado niveles alarmantes con un 227.67% en el índice de precios al consumidor y un 222.56% en el índice de precios al productor. Esta hiperinflación ha erosionado el poder adquisitivo y ha generado incertidumbre económica. El PIB real ha disminuido un 4.16%, reflejando una contracción económica significativa. A pesar de esto, el PIB nominal ha mostrado una leve disminución, situándose en $602.87 mil millones. La tasa de desempleo ha aumentado al 9.15%, lo que indica un deterioro en el mercado laboral. La balanza por cuenta corriente ha mejorado, pasando de un déficit de $20.96 mil millones en 2023 a un superávit de $19.07 mil millones en 2024, lo que sugiere una mejora en el comercio exterior. Sin embargo, la calidad de los activos sigue siendo una preocupación, con una morosidad del 2.76% en la cartera bruta y una cobertura de provisiones del 123.2%. La rentabilidad de los activos (ROA) y del patrimonio (ROE) han mostrado una mejora significativa, con un ROA del 6.55% y un ROE del 65.47%, lo que indica una mayor eficiencia en la generación de utilidades. El sector bancario ha mostrado una mayor solvencia con un índice de solvencia del 31.08% y un Tier 1 del 30.49%. Sin embargo, la intermediación financiera ha disminuido, reflejando una menor actividad crediticia. En términos de liquidez, los activos líquidos más inversiones sobre depósitos han mejorado, situándose en 1.023, lo que indica una mayor capacidad de los bancos para hacer frente a retiros de depósitos. En resumen, Argentina enfrenta desafíos significativos en términos de inflación y crecimiento económico, pero muestra señales de mejora en la balanza por cuenta corriente y la rentabilidad del sector bancario. La estabilidad macroeconómica sigue siendo un reto crucial para el país.",
                    "EvolucionDeMargenesMfinBruto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "EvolucionDeMargenesFecha": [
                        "2019",
                        "2020",
                        "2021",
                        "2022",
                        "2023"
                    ],
                    "EvolucionDeMargenesMfinNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "EvolucionDeMargenesMNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "MorosidadCoberturaMfinBruto": [
                        0.351094,
                        0.502441,
                        0.411269,
                        0.362845,
                        0.292789
                    ],
                    "MorosidadCoberturaFecha": [
                       "2019",
                        "2020",
                        "2021",
                        "2022",
                        "2023"
                    ],
                    "MorosidadCoberturaMfinNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "MorosidadCoberturaMNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "LiquidezFecha": [
                       "2019",
                        "2020",
                        "2021",
                        "2022",
                        "2023"
                    ],
                    "LiquidezMfinBruto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "LiquidezMfinNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "LiquidezMNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "RentabilidadFecha": [
                        "2023",
                        "2022",
                        "2021",
                        "2020",
                        "2019"
                    ],
                    "RentabilidadMfinBruto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "RentabilidadMfinNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "RentabilidadMNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "CapitalizacionFecha": [
                       |"2019",
                        "2020",
                        "2021",
                        "2022",
                        "2023"
                    ],
                    "CapitalizacionMfinBruto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "CapitalizacionMfinNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "CapitalizacionMNeto": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "PanoramaFechas": [
                        "2019",
                        "2020",
                        "2021",
                        "2022",
                        "2023"
                    ],
                    "Balance": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Activo": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Cartera": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Deposito": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                     "Patrimonio": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                     "Patrimonio": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Ingresos (USD MM)": [
                      1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                     "Utilidad (USD MM)": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                     "ROA": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "ROE": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Gastos de transf./Activos": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Gastos de transf./Ingresos Fin.": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Act. Líq + Inv / Depósitos": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Act. Líq. + Inv / Dep. + Ds. Fin": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Intermediación financiera": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Intermediación financiera ampliada": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "LCR": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "NSFR": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Cartera Vencida/Cartera Bruta": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Provisiones/Cartera Vencida": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Provisiones / Cartera Total": [
                       1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Cartera Renegociada / Cartera Bruta": [
                      1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Patrimonio/Activos Totales": [
                      1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Índice de solvencia": [
                      1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Tier 1": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],


                    " / USD": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "VARIACION": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "INGRESOSFINANCIEROS": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Estructura de Activos Activos líquidos": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Estructura de Activos Cartera neta": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Estructura de Activos Cartera Otros activos": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "Estructura de Activos Cartera Activo Total": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "% Activos líquidos": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "%  Cartera neta": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ],
                    "%  Otros activos": [
                        1245,
                        1245,
                        123,
                        124,
                        235
                    ]
                
                }

        """

    Promptcompressed = f"""Por favor, dame un json con el mismo esquema siguiente: {jsonfinal},
    realiza una extracción de información siguiendo estrictamente estas instrucciones:

    Usa únicamente la información dada, sin agregar datos externos.
    Ten en cuenta que las variables pueden estar distribuidas por diferentes años de análisis; identifica y organiza los datos según el año correspondiente.
    La respuesta debe estar en español, en formato JSON, clara y concisa, utilizando siempre comillas dobles.
    Solo responde con el JSON, sin saltos de línea ni texto adicional, y sin la palabra "JSON" al inicio o al final.
    Asegúrate de que el JSON tenga el formato correcto.
    Los campos por extraer están en un array,cada uno con un nombre de campo y una descripción.
    Para los campos que incluyen montos, convierte el monto a un valor numérico sin unidades, si el monto está en formato como 'USD 60 M', convierte el valor a su representación numérica en base 10 por ejemplo USD 60 M a 60000000. En datos numericos nunca incluyas comas para separadores de miles.
    Al brindar la información extraída ten en cuenta el tipo de dato que se indica en la descripción de cada campo.
    Si no encuentras al menos un dato de los campos tipo array, responde con el array vacio nunca iguales a los datos array a null.

    Realiza la extracción en cuenta solo los siguientes documentos: [Documento SyP]: {data_syp}. [Documento SISTEMAS]: {markdown}"""
    
    system_propmt_extraccion ="Eres un asistente especializado en extraer información estructurada de documentos y presentarla en un formato JSON."

    # Obtener la fecha y hora actuales
    fecha_actual = datetime.now()
    id_propmt1 = fecha_actual.strftime("%Y%m%d%H%M%S")
    json_batch_gpt_element_extraccion = {
            "custom_id": f"{id_propmt1}",
            "method": "POST",
            "url": "/chat/completions",
            "body": {"model": "gpt-4o-batch",
                    "messages": [{"role": "system", "content": system_propmt_extraccion},
                                {"role": "user", "content": Promptcompressed}]}}
    
    blob = source_container_client.get_blob_client(f"{anio}/requests_gpt_paises.jsonl").download_blob().readall()

    json_lines = blob.decode("utf-8").splitlines() if blob else []

    # Agregar el nuevo objeto JSON como una línea
    json_lines.append(json.dumps(json_batch_gpt_element_extraccion))

    # Escribir el contenido actualizado en el blob en formato JSONL
    jsonl_content = "\n".join(json_lines) + "\n"

    local_file_path = "requests_gpt_paises.jsonl"
    with open(local_file_path, "w", encoding="utf-8") as file:
        file.write(jsonl_content)
    # Obtener el cliente para el blob específico
    blob_client = source_container_client.get_blob_client(f"{anio}/requests_gpt_paises.jsonl")

    # Subir el contenido actualizado al blob
    blob_client.upload_blob(jsonl_content, overwrite=True)
    

    print("Archivo .jsonl actualizado correctamente.")

    return json_batch_gpt_element_extraccion


def ResultadoGeneral(jsondata):  
    Promptcompressed = f"""A partir de los datos proporcionados ({jsondata}), realiza un análisis financiero general de la region(todos los paises) basandote en los datos proporcionados anteriormente.
    Sigue estrictamente estas instrucciones para tu respuesta:
    Utiliza exclusivamente la información proporcionada sin añadir datos externos.
    La respuesta debe ser clara, concisa y en español 
    No utilices Markdown
    """
    system_propmt_paises ="Eres un asistente experto en el análisis financiero y economico de paises de latino america y el caribe"

    # Obtener la fecha y hora actuales
    fecha_actual = datetime.now()
    id_propmt2 = fecha_actual.strftime("%Y%m%d%H%M%S")
    json_batch_gpt_element_analisis_general= {
            "custom_id": f"{id_propmt2}",
            "method": "POST",
            "url": "/chat/completions",
            "body": {"model": "gpt-4o-batch",
                    "messages": [{"role": "system", "content": system_propmt_paises},
                                {"role": "user", "content": Promptcompressed}]}}
    
    blob = source_container_client.get_blob_client(f"{anio}/requests_gpt_general.jsonl").download_blob().readall()

    json_lines = blob.decode("utf-8").splitlines() if blob else []

    # Agregar el nuevo objeto JSON como una línea
    json_lines.append(json.dumps(json_batch_gpt_element_analisis_general))

    # Escribir el contenido actualizado en el blob en formato JSONL
    jsonl_content = "\n".join(json_lines) + "\n"

    local_file_path = "requests_gpt_general.jsonl"
    with open(local_file_path, "w", encoding="utf-8") as file:
        file.write(jsonl_content)
    # Obtener el cliente para el blob específico
    blob_client = source_container_client.get_blob_client(f"{anio}/requests_gpt_general.jsonl")

    # Subir el contenido actualizado al blob
    blob_client.upload_blob(jsonl_content, overwrite=True)
    
    print("Archivo .jsonl actualizado correctamente.")
    return json_batch_gpt_element_analisis_general

# COMMAND ----------

data_syp=get_document_text(
    file_path=filepathsyp,
    blob_service=blob_service,
    source_container_client=source_container_client,
    formrecognizerservice=formrecognizerservice,
    credentialformrecognizer=credentialformrecognizer
)


# COMMAND ----------

blob = source_container_client.get_blob_client(f"{anio}/ParametrosAnio.json").download_blob().readall()
json_data = json.loads(blob)
procesamientojson = json_data

# COMMAND ----------

paisesaprocesar = procesamientojson.get('Paises')

for pais in paisesaprocesar:
    pdfpath = f"{pais.get('syp')}"
    excelpath = f"{pais.get('excel')}"
    print(f"PDF path : {pdfpath}")
    print(f"Excel path : {excelpath}")
    file = ResultadoPais(excelpath,pdfpath)

# COMMAND ----------

blob = source_container_client.get_blob_client(f"{anio}/requests_gpt_paises.jsonl").download_blob().readall()

json_lines = blob.decode("utf-8").splitlines() if blob else []
jsonl_content = "\n".join(json_lines) + "\n"
local_file_path = "requests_gpt_paises.jsonl"
with open(local_file_path, "w", encoding="utf-8") as file:
    file.write(jsonl_content)


# COMMAND ----------

file = client.files.create(
    file=open("requests_gpt_paises.jsonl", "rb"), 
    purpose="batch"
)
print(file.model_dump_json(indent=2))
file_id = file.id

#Validacion de estado 

status = "pending"
while status != "processed":
    time.sleep(10)
    file_status = client.files.retrieve(file_id)
    status = file_status.status
    print(f"File status: {status}")

    if status not in ["pending", "processed"]:
        raise RuntimeError(f"Unexpected file status: {status}")

# Submit a batch job with the file
batch_response = client.batches.create(
    input_file_id=file_id,
    endpoint="/chat/completions",
    completion_window="24h",
)

# Save batch ID for later use
batch_id = batch_response.id

print(batch_response.model_dump_json(indent=2))

status = "validating"
while status not in ("completed", "failed", "canceled"):
    time.sleep(60)
    batch_response = client.batches.retrieve(batch_id)
    status = batch_response.status
    print(f"{datetime.now()} Batch Id: {batch_id},  Status: {status}")

if batch_response.status == "failed":
    for error in batch_response.errors.data:  
        print(f"Error code {error.code} Message {error.message}")

if batch_response.status == "completed":
    import json

    output_file_id = batch_response.output_file_id

    if not output_file_id:
        output_file_id = batch_response.error_file_id

    if output_file_id:
        file_response = client.files.content(output_file_id)
        raw_responses = file_response.text.strip().split('\n')  
        formatted_json = [json.loads(raw_response) for raw_response in raw_responses]  # Parse into a list of dictionaries
        print("jsooooon")
        print(formatted_json)




# COMMAND ----------

# Credenciales de Cosmos DB
endpoint = 'https://cd-poc-ifis-cr.documents.azure.com:443/'
key = 'ilmNXXg4kDypH4hJyqNBXRr2HUSS3AlWHuhVFrzmQ2YCJf0CqrYqzXzXAsrsZwvvorN3DiAaljvSACDbbDv6MQ=='
database_name = 'IfisCAF'
container_name = 'Reportes'


# COMMAND ----------

clientCOSMOS = CosmosClient(endpoint, key)
database = clientCOSMOS.get_database_client(database_name)
container = database.get_container_client(container_name)


# COMMAND ----------

# Lista para almacenar los resultados
results = []
PaisesResultado = []
if formatted_json:
    for entry in formatted_json:
        custom_id = entry.get('custom_id')
        response_body = entry.get('response', {}).get('body', {})
        content = None

        # Navegar hasta el contenido del JSON si está disponible
        if 'choices' in response_body:
            content = response_body['choices'][0]['message']['content']
        
        # Limpiar el contenido para extraer el JSON real (si es necesario)
        if content and content.startswith('```json'):
            content = content.strip('```json').strip()

        # Almacenar el resultado en la lista
        results.append({'custom_id': custom_id, 'content': content})

    print("Resultados antes de guardar el item:" results)

    # Mostrar resultados almacenados
    for result in results:
        print(f"Custom ID: {result['custom_id']}")
        PaisesResultado.append(json.loads(result['content']))
        print(f"Content: {result['content']}\n")
    # Consulta SQL para buscar el elemento por "AÑO DE CARGUE"
    query = f"SELECT * FROM c WHERE c['AÑO DE CARGUE'] = {anio}"

    # Ejecutar la consulta
    items = list(container.query_items(query=query, enable_cross_partition_query=True))

    if items:
        existing_item = items[0]  # Tomar el primer resultado (si hay más de uno, debes manejarlo)
        existing_item["status"] = "Procesando contexto general"  # Actualiza o añade campos
        existing_item["Paises"] = PaisesResultado
        existing_item["CONTEXTO"] = "En construcción"
        container.replace_item(item=existing_item, body=existing_item)
    else:
        print("No se encontró el elemento.")




# COMMAND ----------

# MAGIC %md
# MAGIC ###Renombrar los json de parametros

# COMMAND ----------


# Obtener el cliente del blob original
blob_original = source_container_client.get_blob_client(f"{anio}/ParametrosAnio.json")

# Generar el nuevo nombre con fecha y hora actual
fecha_hora_actual = datetime.now().strftime("%Y%m%d%H%M%S")
nuevo_nombre = f"{anio}/ParametrosAnio-{fecha_hora_actual}.json"

# Obtener el cliente del blob nuevo
blob_nuevo = source_container_client.get_blob_client(nuevo_nombre)

# Copiar el contenido del blob original al nuevo blob
blob_nuevo.start_copy_from_url(blob_original.url)

# Confirmar que la copia se completó y luego eliminar el blob original
blob_original.delete_blob()

# COMMAND ----------

# Obtener el cliente del blob original
blob_jsnol = source_container_client.get_blob_client(f"{anio}/requests_gpt_paises.jsonl")

nuevo_nombrejsonl = f"{anio}/requests_gpt_paises-{fecha_hora_actual}.jsonl"

# Obtener el cliente del blob nuevo
blob_nuevo_jsonl = source_container_client.get_blob_client(nuevo_nombrejsonl)

# Copiar el contenido del blob original al nuevo blob
blob_nuevo_jsonl.start_copy_from_url(blob_jsnol.url)

# Confirmar que la copia se completó y luego eliminar el blob original
blob_jsnol.delete_blob()

# COMMAND ----------

# MAGIC %md
# MAGIC # Contexto general

# COMMAND ----------

resultadogeneraljson = ResultadoGeneral(PaisesResultado)
blob = source_container_client.get_blob_client(f"{anio}/requests_gpt_general.jsonl").download_blob().readall()

json_lines = blob.decode("utf-8").splitlines() if blob else []
jsonl_content = "\n".join(json_lines) + "\n"
local_file_path = "requests_gpt_general.jsonl"
with open(local_file_path, "w", encoding="utf-8") as file:
    file.write(jsonl_content)

# COMMAND ----------

file = client.files.create(
    file=open("requests_gpt_general.jsonl", "rb"), 
    purpose="batch"
)
print(file.model_dump_json(indent=2))
file_id = file.id

#Validacion de estado 

status = "pending"
while status != "processed":
    time.sleep(10)
    file_status = client.files.retrieve(file_id)
    status = file_status.status
    print(f"File status: {status}")

    if status not in ["pending", "processed"]:
        raise RuntimeError(f"Unexpected file status: {status}")

# Submit a batch job with the file
batch_response = client.batches.create(
    input_file_id=file_id,
    endpoint="/chat/completions",
    completion_window="24h",
)

# Save batch ID for later use
batch_id = batch_response.id

print(batch_response.model_dump_json(indent=2))

status = "validating"
while status not in ("completed", "failed", "canceled"):
    time.sleep(60)
    batch_response = client.batches.retrieve(batch_id)
    status = batch_response.status
    print(f"{datetime.now()} Batch Id: {batch_id},  Status: {status}")

if batch_response.status == "failed":
    for error in batch_response.errors.data:  
        print(f"Error code {error.code} Message {error.message}")

if batch_response.status == "completed":
    import json

    output_file_id = batch_response.output_file_id

    if not output_file_id:
        output_file_id = batch_response.error_file_id

    if output_file_id:
        file_response = client.files.content(output_file_id)
        raw_responses = file_response.text.strip().split('\n')  
        formatted_json = [json.loads(raw_response) for raw_response in raw_responses]  # Parse into a list of dictionaries
        print("jsooooon")
        print(formatted_json)

# COMMAND ----------

resultadocontextognrl = ""
if formatted_json:
    for entry in formatted_json:
        custom_id = entry.get('custom_id')
        response_body = entry.get('response', {}).get('body', {})
        content = None

        # Navegar hasta el contenido del JSON si está disponible
        if 'choices' in response_body:
            content = response_body['choices'][0]['message']['content']
        
        # Limpiar el contenido para extraer el JSON real (si es necesario)
        if content and content.startswith('```json'):
            content = content.strip('```json').strip()

        # Almacenar el resultado en la lista
        resultadocontextognrl = content
    
    # Consulta SQL para buscar el elemento por "AÑO DE CARGUE"
    query = f"SELECT * FROM c WHERE c['AÑO DE CARGUE'] = {anio}"

    # Ejecutar la consulta
    items = list(container.query_items(query=query, enable_cross_partition_query=True))

    if items:
        existing_item = items[0] 
        existing_item["CONTEXTO"] = resultadocontextognrl
        existing_item["status"] = "Proceso completado"
        container.replace_item(item=existing_item, body=existing_item)
    else:
        print("No se encontró el elemento.")


# COMMAND ----------

# Obtener el cliente del blob original
blob_jsnol = source_container_client.get_blob_client(f"{anio}/requests_gpt_general.jsonl")

nuevo_nombrejsonl = f"{anio}/requests_gpt_general-{fecha_hora_actual}.jsonl"

# Obtener el cliente del blob nuevo
blob_nuevo_jsonl = source_container_client.get_blob_client(nuevo_nombrejsonl)

# Copiar el contenido del blob original al nuevo blob
blob_nuevo_jsonl.start_copy_from_url(blob_jsnol.url)

# Confirmar que la copia se completó y luego eliminar el blob original
blob_jsnol.delete_blob()
